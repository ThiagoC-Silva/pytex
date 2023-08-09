from openpyxl import Workbook
import os


class WorkbookGenerator:
    def __init__(self, header):
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = 'Report'
        self.create_header(header)


    def create_header(self, header):
        for col, val in enumerate(header, start = 1):
            self.worksheet.cell(row = 1, column = col, value = val)
        self.save_reports_in_folder()

    
    def insertion_cells(self,data_row, status, line):
        if status == False:
            for col, val in enumerate(data_row.values() , start = 1):
                self.worksheet.cell(row = line, column = col, value = val )
        else:
            add_weight = len(data_row) + 1
            self.worksheet.cell(row = 1, column = add_weight, value = 'Weight')
            self.worksheet.cell(row = line, column = add_weight, value = data_row['Weight'])

        self.save_reports_in_folder()


    def save_reports_in_folder(self):
        reports = 'reports/'
        if not os.path.exists(reports):
            os.mkdir(reports)
        
        file_name = 'report.xlsx'
        path_folder = reports + file_name
        self.workbook.save(path_folder)
