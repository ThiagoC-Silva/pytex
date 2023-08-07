#Biblioteca para lidar com dados JSON
import json
#Bibliotexa para a criação e manipulação de planilhas no Excel
from openpyxl import Workbook


file_excel = Workbook() #Cria um arquivo excel em branco 
spreadsheet = file_excel.active #Acesso a planilha recém criada
spreadsheet.title = 'Title' #Título da tabela

# Abrir o arquivo JSON em modo de leitura e amazenar os dados na vairável 'file'
with open('data.json', 'r') as file:
    object = json.load(file) # String JSON para objeto

linha = 1
coluna = 1
vPlacaAtual = ""
vTipoAtual = ""
vTipoAtualFor = ""
vPlacaAtualFor = ""

#Cabeçalho da planilha
spreadsheet.cell(row=1, column=1, value="Day")
spreadsheet.cell(row=1, column=2, value="Plate")
spreadsheet.cell(row=1, column=3, value="Vehicle")
spreadsheet.cell(row=1, column=4, value="Day Period")


for k in object['rows']:    
    
    if (vTipoAtual == ""):
        vTipoAtualFor = k["vTipo"]
    else :
        vTipoAtualFor = vTipoAtual

    if (vPlacaAtual == ""):
        vPlacaAtualFor = k["vPlaca"]
    else :
        vPlacaAtualFor = vPlacaAtual        

    vTipoAtual = k["vTipo"]

    if (vPlacaAtual != k["vPlaca"]):        		
        vPlacaAtual = k["vPlaca"]	         
        vTipoAtual = k["vTipo"]
        
        linha = linha + 1 
        
        spreadsheet.cell(row=linha, column=1, value=k["dia"])
        spreadsheet.cell(row=linha, column=2, value=k["vPlaca"])
        spreadsheet.cell(row=linha, column=3, value=k["vDescricao_veiculo"])
        spreadsheet.cell(row=linha, column=4, value=k["vTipo"])   
        coluna = 5     
 
    if (vTipoAtualFor != vTipoAtual) and (vPlacaAtualFor == vPlacaAtual):
        linha = linha + 1

        spreadsheet.cell(row=linha, column=1, value=k["dia"])
        spreadsheet.cell(row=linha, column=2, value=k["vPlaca"])
        spreadsheet.cell(row=linha, column=3, value=k["vDescricao_veiculo"])
        spreadsheet.cell(row=linha, column=4, value=k["vTipo"])     

        coluna = 5       

    spreadsheet.cell(row=1, column=coluna, value="Peso")  
    spreadsheet.cell(row=linha, column=coluna, value=k["nPeso"])  
    coluna = coluna + 1

    print(vPlacaAtualFor, vPlacaAtual, "-" ,vTipoAtualFor, vTipoAtual)


print(file_excel.sheetnames)
file_excel.save("relatorio.xlsx")