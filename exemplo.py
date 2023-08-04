#deixar o jason em um arquivo externo
#Script faz a leitura
#traduzir para o inglês

#Biblioteca para trabalhar com dados no formato em JASON
import json

#Biblioteca para criar e manipular planilhas no Excel
from openpyxl import Workbook


arquivo_excel = Workbook() # Cria um novo arquivo excel
planilha1 = arquivo_excel.active # Obtém a planilha recém criada
planilha1.title = "title" # Título da planilha

texto = """
{
	"rows":
	[
		{
			"dia": 1,
			"vPlaca": "ABC-1231",
			"vDescricao_veiculo": "CAM. COMPAC. VOLKS",
			"nPeso": 2.49,
			"vTipo": "DIURNO"
		},
		{
			"dia": 2,
			"vPlaca": "CBA-1321",
			"vDescricao_veiculo": "CAM. COMPAC. VOLKS 17.280",
			"nPeso": 3.01,
			"vTipo": "DIURNO"
		},
		{
			"dia": 3,
			"vPlaca": "LOP-7894",
			"vDescricao_veiculo": "CAM. COMPAC. VOLKS 17.281",
			"nPeso": 6.42,
			"vTipo": "NOTURNO"
		},
		{
			"dia": 3,
			"vPlaca": "LOP-7894",
			"vDescricao_veiculo": "CAM. COMPAC. VOLKS 17.281",
			"nPeso": 8.89,
			"vTipo": "NOTURNO"
		},
		{
			"dia": 3,
			"vPlaca": "LOP-7894",
			"vDescricao_veiculo": "CAM. COMPAC. VOLKS 17.281",
			"nPeso": 5.15,
			"vTipo": "DIURNO"
		}
	]
}
"""

obj = json.loads(texto) #String JSON para objeto Python

#Variáveis que serão usadas para rastrear os valores 
linha = 1
coluna = 1
vPlacaAtual = ""
vTipoAtual = ""
vTipoAtualFor = ""
vPlacaAtualFor = ""

#Cabçalho da planilha
planilha1.cell(row=1, column=1, value="Dia")
planilha1.cell(row=1, column=2, value="Placa")
planilha1.cell(row=1, column=3, value="Veículo")
planilha1.cell(row=1, column=4, value="Tipo")


for k in obj['rows']:    
    
    if (vTipoAtual == ""):
        vTipoAtualFor = k["vTipo"]
    else :
        vTipoAtualFor = vTipoAtual

    if (vPlacaAtual == ""):
        vPlacaAtualFor = k["vPlaca"]
    else :
        vPlacaAtualFor = vPlacaAtual        

    vTipoAtual = k["vTipo"]

    if (vPlacaAtual != k["vPlaca"]):        		
        vPlacaAtual = k["vPlaca"]	         
        vTipoAtual = k["vTipo"]
        
        linha = linha + 1 
        
        planilha1.cell(row=linha, column=1, value=k["dia"])
        planilha1.cell(row=linha, column=2, value=k["vPlaca"])
        planilha1.cell(row=linha, column=3, value=k["vDescricao_veiculo"])
        planilha1.cell(row=linha, column=4, value=k["vTipo"])   
        coluna = 5     
 
    if (vTipoAtualFor != vTipoAtual) and (vPlacaAtualFor == vPlacaAtual):
        linha = linha + 1

        planilha1.cell(row=linha, column=1, value=k["dia"])
        planilha1.cell(row=linha, column=2, value=k["vPlaca"])
        planilha1.cell(row=linha, column=3, value=k["vDescricao_veiculo"])
        planilha1.cell(row=linha, column=4, value=k["vTipo"])     

        coluna = 5       

    planilha1.cell(row=1, column=coluna, value="Peso")  
    planilha1.cell(row=linha, column=coluna, value=k["nPeso"])  
    coluna = coluna + 1

    print(vPlacaAtualFor, vPlacaAtual, "-" ,vTipoAtualFor, vTipoAtual)


print(arquivo_excel.sheetnames)
arquivo_excel.save("relatorio.xlsx")